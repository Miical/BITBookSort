from openpyxl import *
from openpyxl.styles import PatternFill
from idx_parser import parser
from compare import bookidx_cmp
from functools import cmp_to_key

def load_file():
    while True:
        file_path = input("请输入书单表格路径: ")

        try:
            workbook = load_workbook(file_path)
            _ = workbook.sheetnames
            print(f"文件 {file_path} 加载成功。")
            return workbook
        except FileNotFoundError:
            print("文件不存在，请检查路径是否正确。")
        except Exception as e:
            print("错误:", e)

def sort_by_book_idx(booklist, columns_idx):
    def key_func(item1, item2):
        return bookidx_cmp(item1[columns_idx], item2[columns_idx])
    return sorted(booklist, key=cmp_to_key(key_func))

if __name__ == '__main__':
    workbook = load_file()
    sheet_names = workbook.sheetnames

    # Select Sheet
    print("可用的sheet:")
    for i, name in enumerate(sheet_names):
        print(f"{i+1}. {name}")
    selected_sheet_index = int(input("请选择一个sheet（输入编号）: ")) - 1

    if selected_sheet_index < 0 or selected_sheet_index >= len(sheet_names):
        print("选择的sheet编号无效，请重新运行程序并输入正确的编号。")
        input("按回车键退出...")
        exit(1)

    selected_sheet_name = sheet_names[selected_sheet_index]
    sheet = workbook[selected_sheet_name]

    # Select Column
    columns_info = [cell.value for cell in sheet[1]]

    print(f"{selected_sheet_name} 表格列信息如下:")
    for i, column_name in enumerate(columns_info):
        print(f"{i+1}. {column_name}")

    selected_column_index = int(input("请选择含有图书索引号的列（输入编号）: ")) - 1

    if selected_column_index < 0 or selected_column_index >= len(columns_info):
        print("选择的列编号无效，请重新运行程序并输入正确的编号。")
        input("按回车键退出...")
        exit(0)

    # Parse and Sort
    parse_success = [row for row in sheet.iter_rows(min_row=2, values_only=True) \
                    if parser(row[selected_column_index]) is not None]
    parse_fail = [row for row in sheet.iter_rows(min_row=2, values_only=True) \
                    if parser(row[selected_column_index]) is None]

    if len(parse_fail) > 0:
        print("以下书目索引号无法解析（无法解析的书目将被放置在表格尾部）:")
        for row in parse_fail:
            print(row)
        print()

    sorted_data = sort_by_book_idx(parse_success, selected_column_index)
    parse_fail = sorted(parse_fail, key=lambda x: x[selected_column_index])

    # Create new workbook and sheet
    new_workbook = Workbook()
    new_sheet = new_workbook.active

    new_sheet.append(columns_info)

    for row in sorted_data + parse_fail:
        new_sheet.append(row)

    # Highlight failed rows
    fail_num = len(parse_fail)
    if fail_num > 0:
        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        for row in new_sheet.iter_rows(max_row=new_sheet.max_row, min_row=max(new_sheet.max_row-fail_num+1, 2)):
            for cell in row:
                cell.fill = red_fill

    # Save new workbook
    new_file_path = 'sorted_booklist.xlsx'
    new_workbook.save(new_file_path)
    print(f"排序后的新表格已保存至 {new_file_path}。")
    input("按回车键退出...")
